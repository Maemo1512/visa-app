from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Font
import sqlite3
from datetime import datetime

app = Flask(__name__)
app.config['JSONIFY_PRETTYPRINT_REGULAR'] = False
CORS(app, resources={r"/*": {"origins": "*"}})

DATABASE = "database.db"

# CREATE DATABASE
def init_db():
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS expenses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item TEXT,
            qty INTEGER,
            unit_cost REAL,
            total REAL,
            date TEXT
        )
    """)

    conn.commit()
    conn.close()

init_db()

# ADD EXPENSE
@app.route('/add-expense', methods=['POST'])
def add_expense():
    data = request.json

    item = data['item']
    qty = data['qty']
    unit_cost = data['unit_cost']
    total = qty * unit_cost
    date = datetime.now().strftime("%Y-%m-%d")

    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO expenses (item, qty, unit_cost, total, date)
        VALUES (?, ?, ?, ?, ?)
    """, (item, qty, unit_cost, total, date))

    conn.commit()
    conn.close()

    return jsonify({
        "message": "Expense added successfully"
    })

# GET EXPENSES
@app.route('/expenses', methods=['GET'])
def get_expenses():

    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM expenses")
    rows = cursor.fetchall()

    conn.close()

    expenses = []

    for row in rows:
        expenses.append({
            "id": row[0],
            "item": row[1],
            "qty": row[2],
            "unit_cost": row[3],
            "total": row[4],
            "date": row[5]
        })

    return jsonify(expenses)
# DELETE EXPENSE
@app.route('/delete-expense/<int:id>', methods=['DELETE'])
def delete_expense(id):

    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()

    cursor.execute("DELETE FROM expenses WHERE id = ?", (id,))

    conn.commit()
    conn.close()

    return jsonify({
        "message": "Expense deleted"
    })
# UPDATE EXPENSE
@app.route('/update-expense/<int:id>', methods=['PUT'])
def update_expense(id):

    data = request.json

    item = data['item']
    qty = data['qty']
    unit_cost = data['unit_cost']
    total = qty * unit_cost

    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE expenses
        SET item = ?, qty = ?, unit_cost = ?, total = ?
        WHERE id = ?
    """, (item, qty, unit_cost, total, id))

    conn.commit()
    conn.close()

    return jsonify({
        "message": "Expense updated"
    })

# EXPORT EXCEL
@app.route('/export-excel', methods=['GET'])
def export_excel():

    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM expenses")
    rows = cursor.fetchall()

    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Expense Report"

    headers = ["ID", "Item", "Quantity", "Unit Cost", "Total", "Date"]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    for row_num, row_data in enumerate(rows, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num).value = value

    file_name = "expense_report.xlsx"
    wb.save(file_name)

    return send_file(file_name, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)